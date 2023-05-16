from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import os, zipfile, shutil, pickle


def df_to_xlsx(df, name):
    wb = Workbook()
    ws = wb.active
    ws.title = "search_result"
    for r in dataframe_to_rows(df, index=False, header=True):
        try:
            ws.append(r)
        except:
            print("jump", r[0])
    column_widths = []
    first_row = True
    for row in ws:
        for i, cell in enumerate(row):
            if first_row:
                size = cell.font.size
                cell.font = Font(bold=True, size=size + 3)
            else:
                pass
            cell.alignment = Alignment(
                wrapText=True, horizontal="left", vertical="center"
            )
            if len(column_widths) > i:
                if len(cell.value) > column_widths[i]:
                    column_widths[i] = len(cell.value)
            else:
                column_widths += [len(cell.value)]
        first_row = False

    # arbitrarily set to 500
    column_widths[2] = 500

    for i, column_width in enumerate(column_widths, 1):  # ,1 to start at 1
        ws.column_dimensions[get_column_letter(i)].width = column_width

    wb.save(name)


def create_result_folder(folder_name):
    try:
        os.mkdir(folder_name)
        print("Current dir in : " + os.getcwd())
    except:
        shutil.rmtree(folder_name)
        print(f"{folder_name} folder exist, removed existing {folder_name} folder")
        os.mkdir(folder_name)
        print(f"Created {folder_name} folder")


def zip_txt_files(folder, zfile):
    with zipfile.ZipFile(f"./{folder}/{zfile}", "w") as zf:
        for file in os.listdir(folder):
            if file.endswith(".txt"):
                path = os.path.join(folder, file)
                zf.write(path)


def store_var(dictionary, file_name):
    path = os.path.join("pickles", file_name)
    with open(path, "wb") as f:
        pickle.dump(dictionary, f)


def get_var(file_name):
    path = os.path.join("pickles", file_name)
    with open(path, "rb") as f:
        loaded_dict = pickle.load(f)
    return loaded_dict


def back_search(dictionary):
    reverse = {}
    for k, v in dictionary.items():
        if v not in reverse:
            reverse[v] = k
        else:
            reverse[v].append(k)
    return reverse
